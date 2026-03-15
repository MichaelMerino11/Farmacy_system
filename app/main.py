from fastapi import FastAPI, Request, Form, HTTPException, Query
from fastapi.responses import HTMLResponse, RedirectResponse, Response, JSONResponse
from fastapi.templating import Jinja2Templates
from fastapi.staticfiles import StaticFiles
import os
import platform
import re
from datetime import datetime

from sqlalchemy.orm import Session
from .db import Base, engine, SessionLocal
from .models import Muestra, Caja

import barcode
from barcode.writer import ImageWriter

from weasyprint import HTML as WeasyHTML
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
import io
import base64


app = FastAPI(title="UTN - Laboratorio")

os.makedirs("app/static", exist_ok=True)
os.makedirs("app/static/barcodes", exist_ok=True)
os.makedirs("app/media", exist_ok=True)

app.mount("/static", StaticFiles(directory="app/static"), name="static")
app.mount("/media",  StaticFiles(directory="app/media"),  name="media")
templates = Jinja2Templates(directory="app/templates")
Base.metadata.create_all(bind=engine)

CAPACIDAD_CAJA = 81  # 9 × 9


def get_logo_b64() -> str:
    """Lee utn_logo.png y lo devuelve como data URI base64 para embeber en PDFs."""
    logo_path = os.path.join("app", "media", "utn_logo.png")
    try:
        with open(logo_path, "rb") as f:
            data = base64.b64encode(f.read()).decode("utf-8")
        return f"data:image/png;base64,{data}"
    except Exception:
        return ""  # Si falla, el PDF usará solo texto


# ── UTILIDADES ─────────────────────────────────────────────────────────────────

def contar_muestras_en_caja(db: Session, caja_id: int) -> int:
    return db.query(Muestra).filter(Muestra.caja_id == caja_id).count()


def calcular_ubicacion(caja: Caja, replica: int, tubo: int) -> str:
    tipo = "Vertical" if caja.congelador == 1 else "Horizontal"
    return (
        f"Congelador {tipo} / Piso {caja.piso} / "
        f"Posición {caja.posicion} / Réplica {replica} / Tubo {tubo}"
    )


def generar_codigo_para_caja(caja: Caja, especie: str, origen: str) -> str:
    """
    Formato: {V|H}{piso}-{posicion}-{E|e}-{ORIG4}
    Ejemplo: V1-3-E-UVIL
    """
    tipo   = "V" if caja.congelador == 1 else "H"
    esp    = "E" if especie == "SI" else "e"
    origen_code = (origen or "XX")[:4].upper().replace(" ", "")
    return f"{tipo}{caja.piso}-{caja.posicion}-{esp}-{origen_code}"


def validar_texto(valor, campo, max_len=100, patron=r"^[\w\s\-\.\/áéíóúÁÉÍÓÚñÑ]+$"):
    valor = valor.strip()
    if not valor:
        raise HTTPException(status_code=422, detail=f"El campo '{campo}' no puede estar vacío.")
    if len(valor) > max_len:
        raise HTTPException(status_code=422, detail=f"El campo '{campo}' excede {max_len} caracteres.")
    if not re.match(patron, valor):
        raise HTTPException(status_code=422, detail=f"El campo '{campo}' contiene caracteres no permitidos.")
    return valor


def enviar_a_impresora(data: bytes):
    if platform.system() == "Windows":
        try:
            import win32print
            printer_name = os.getenv("PRINTER_NAME", None)
            if not printer_name:
                printer_name = win32print.GetDefaultPrinter()
            handle = win32print.OpenPrinter(printer_name)
            try:
                win32print.StartDocPrinter(handle, 1, ("Etiqueta UTN", None, "RAW"))
                win32print.StartPagePrinter(handle)
                win32print.WritePrinter(handle, data)
                win32print.EndPagePrinter(handle)
            finally:
                win32print.EndDocPrinter(handle)
                win32print.ClosePrinter(handle)
        except ImportError:
            raise RuntimeError("Falta instalar pywin32. Ejecuta: pip install pywin32")
    else:
        printer_path = os.getenv("PRINTER_PATH", "/dev/usb/lp0")
        with open(printer_path, "wb") as printer:
            printer.write(data)


# ── PÁGINAS ────────────────────────────────────────────────────────────────────

@app.get("/", response_class=HTMLResponse)
def home(request: Request):
    return templates.TemplateResponse("index.html", {"request": request})


@app.get("/dashboard", response_class=HTMLResponse)
def dashboard(request: Request):
    db       = SessionLocal()
    muestras = db.query(Muestra).order_by(Muestra.created_at.desc()).all()
    cajas    = db.query(Caja).order_by(Caja.created_at.desc()).all()
    ocupacion = {c.id: contar_muestras_en_caja(db, c.id) for c in cajas}
    db.close()
    return templates.TemplateResponse("dashboard.html", {
        "request":      request,
        "muestras":     muestras,
        "cajas":        cajas,
        "ocupacion":    ocupacion,
        "capacidad_max": CAPACIDAD_CAJA,
    })


# ── CAJAS — API ────────────────────────────────────────────────────────────────

@app.get("/cajas/disponibles")
def cajas_disponibles():
    """Lista todas las cajas con ocupación actual. Usado por el formulario de muestras."""
    db    = SessionLocal()
    cajas = db.query(Caja).order_by(Caja.id).all()
    resultado = []
    for c in cajas:
        ocupadas = contar_muestras_en_caja(db, c.id)
        libres   = CAPACIDAD_CAJA - ocupadas
        tipo     = "Vertical" if c.congelador == 1 else "Horizontal"
        resultado.append({
            "id":       c.id,
            "nombre":   c.nombre,
            "tipo":     tipo,
            "piso":     c.piso,
            "posicion": c.posicion,
            "fecha":    c.fecha,
            "ocupadas": ocupadas,
            "libres":   libres,
            "llena":    libres <= 0,
            "label":    f"{c.nombre} — {tipo}, Piso {c.piso}, Pos. {c.posicion} ({libres} libres)",
        })
    db.close()
    return resultado


# ── CAJAS — CRUD ───────────────────────────────────────────────────────────────

@app.post("/cajas")
def crear_caja(
    congelador: int = Form(...),
    posicion:   int = Form(...),
    piso:       int = Form(1),
    fecha:      str = Form(...),
    nombre:     str = Form(...),
):
    if congelador not in [1, 2]:
        raise HTTPException(status_code=422, detail="Congelador debe ser 1 (vertical) o 2 (horizontal).")
    if posicion < 1:
        raise HTTPException(status_code=422, detail="La posición debe ser mayor a 0.")
    if piso < 1:
        raise HTTPException(status_code=422, detail="El piso debe ser mayor a 0.")

    nombre = validar_texto(nombre, "Nombre de caja", max_len=100)

    if not re.match(r"^\d{4}-\d{2}-\d{2}$", fecha):
        raise HTTPException(status_code=422, detail="La fecha debe tener formato YYYY-MM-DD.")

    db: Session = SessionLocal()
    existe = db.query(Caja).filter(
        Caja.congelador == congelador,
        Caja.piso       == piso,
        Caja.posicion   == posicion,
    ).first()
    if existe:
        db.close()
        tipo = "vertical" if congelador == 1 else "horizontal"
        raise HTTPException(
            status_code=422,
            detail=f"Ya existe una caja en el congelador {tipo}, piso {piso}, posición {posicion}."
        )

    caja = Caja(congelador=congelador, posicion=posicion, piso=piso, fecha=fecha, nombre=nombre)
    db.add(caja)
    db.commit()
    db.close()
    return RedirectResponse(url="/dashboard?tab=cajas", status_code=303)


@app.get("/cajas/{caja_id}/eliminar")
def eliminar_caja(caja_id: int):
    db   = SessionLocal()
    caja = db.query(Caja).filter(Caja.id == caja_id).first()
    if not caja:
        db.close()
        raise HTTPException(status_code=404, detail="Caja no encontrada.")
    ocupadas = contar_muestras_en_caja(db, caja_id)
    if ocupadas > 0:
        db.close()
        raise HTTPException(
            status_code=422,
            detail=f"No se puede eliminar: la caja tiene {ocupadas} muestra(s) asignada(s)."
        )
    db.delete(caja)
    db.commit()
    db.close()
    return RedirectResponse(url="/dashboard?tab=cajas", status_code=303)


# ── MUESTRAS ───────────────────────────────────────────────────────────────────

@app.post("/muestras")
def crear_muestra(
    caja_id:                   int = Form(...),
    codigo_utn_especie:        str = Form(...),
    numero_replica:            int = Form(...),
    numero_tubo_en_caja:       int = Form(...),
    numero_muestra_ccmbi_ogem: str = Form(...),
    medio_cultivo:             str = Form(...),
    especie:                   str = Form("NO"),
    seguimiento:               str = Form("NO"),
    identificacion_taxonomica: str = Form(""),
    origen_muestra:            str = Form(...),
):
    codigo_utn_especie        = validar_texto(codigo_utn_especie, "Código UTN especie", max_len=50)
    numero_muestra_ccmbi_ogem = validar_texto(numero_muestra_ccmbi_ogem, "N° muestra CCMBIOGEM", max_len=50)
    medio_cultivo             = validar_texto(medio_cultivo, "Medio de cultivo", max_len=50)
    origen_muestra            = validar_texto(origen_muestra, "Origen de la muestra", max_len=100)

    if especie not in ("SI", "NO"):
        especie = "NO"
    if seguimiento not in ("SI", "NO"):
        seguimiento = "NO"
    if numero_replica < 1:
        raise HTTPException(status_code=422, detail="El número de réplica debe ser mayor a 0.")
    if not (1 <= numero_tubo_en_caja <= CAPACIDAD_CAJA):
        raise HTTPException(
            status_code=422,
            detail=f"El tubo debe estar entre 1 y {CAPACIDAD_CAJA}."
        )

    db: Session = SessionLocal()

    # Verificar que la caja existe
    caja = db.query(Caja).filter(Caja.id == caja_id).first()
    if not caja:
        db.close()
        raise HTTPException(status_code=422, detail="La caja seleccionada no existe.")

    # Validar capacidad 81
    ocupadas = contar_muestras_en_caja(db, caja_id)
    if ocupadas >= CAPACIDAD_CAJA:
        # Buscar sugerencia
        todas = db.query(Caja).order_by(Caja.id).all()
        sugerencia = next(
            (c for c in todas if contar_muestras_en_caja(db, c.id) < CAPACIDAD_CAJA),
            None
        )
        db.close()
        msg = (
            f"La caja '{caja.nombre}' está llena "
            f"({CAPACIDAD_CAJA}/{CAPACIDAD_CAJA} muestras). "
            f"Selecciona otra caja o crea una nueva."
        )
        if sugerencia:
            libres = CAPACIDAD_CAJA - contar_muestras_en_caja(SessionLocal(), sugerencia.id)
            msg += f" Sugerencia: '{sugerencia.nombre}' tiene {libres} espacio(s) libre(s)."
        raise HTTPException(status_code=422, detail=msg)

    # Validar que el tubo no esté ocupado en esa caja
    tubo_ocupado = db.query(Muestra).filter(
        Muestra.caja_id             == caja_id,
        Muestra.numero_tubo_en_caja == numero_tubo_en_caja,
    ).first()
    if tubo_ocupado:
        db.close()
        raise HTTPException(
            status_code=422,
            detail=f"El tubo #{numero_tubo_en_caja} ya está ocupado en la caja '{caja.nombre}'."
        )

    # Validar CCMBIOGEM único
    existe = db.query(Muestra).filter(
        Muestra.numero_muestra_ccmbi_ogem == numero_muestra_ccmbi_ogem
    ).first()
    if existe:
        db.close()
        raise HTTPException(
            status_code=422,
            detail=f"Ya existe una muestra con N° CCMBIOGEM '{numero_muestra_ccmbi_ogem}'."
        )

    # Generar código de barras correlativo
    ultimo       = db.query(Muestra).order_by(Muestra.id.desc()).first()
    nuevo_numero = 1 if not ultimo else ultimo.id + 1
    codigo_barra = f"UTN-2026-{str(nuevo_numero).zfill(5)}"

    numero_caja_label = f"CAJA-{str(caja.id).zfill(3)}"
    ubicacion         = calcular_ubicacion(caja, numero_replica, numero_tubo_en_caja)
    codigo_caja       = generar_codigo_para_caja(caja, especie, origen_muestra)

    muestra = Muestra(
        caja_id                   = caja_id,
        numero_caja               = numero_caja_label,
        nivel                     = f"Piso {caja.piso}",
        codigo_utn_especie        = codigo_utn_especie,
        numero_replica            = numero_replica,
        numero_tubo_en_caja       = numero_tubo_en_caja,
        numero_muestra_ccmbi_ogem = numero_muestra_ccmbi_ogem,
        medio_cultivo             = medio_cultivo,
        ubicacion_refrigerador    = ubicacion,
        codigo_barra              = codigo_barra,
        especie                   = especie,
        seguimiento               = seguimiento,
        identificacion_taxonomica = identificacion_taxonomica.strip(),
        origen_muestra            = origen_muestra,
        codigo_para_caja          = codigo_caja,
    )
    db.add(muestra)
    db.commit()
    db.refresh(muestra)

    try:
        code128 = barcode.get("code128", muestra.codigo_barra, writer=ImageWriter())
        code128.save(f"app/static/barcodes/{muestra.codigo_barra}")
    except Exception:
        pass  # no bloquear si falla el barcode

    db.close()
    return RedirectResponse(url=f"/?muestra_id={muestra.id}", status_code=303)


@app.get("/muestras/{muestra_id}/print-raw")
def imprimir_raw(muestra_id: int):
    db      = SessionLocal()
    muestra = db.query(Muestra).filter(Muestra.id == muestra_id).first()
    db.close()
    if not muestra:
        return {"error": "No existe"}

    codigo_barcode = str(muestra.id).zfill(5)
    codigo_texto   = f"UTN-2026-{codigo_barcode}"

    data  = b'\x1b\x40'
    data += b'\x1b\x61\x02'
    data += b'\x1d\x68\x28'
    data += b'\x1d\x77\x02'
    data += b'\x1d\x48\x00'
    data += b'\x1d\x6b\x49'
    data += bytes([len(codigo_barcode.encode())])
    data += codigo_barcode.encode()
    data += b'\x1b\x4a\x04'
    data += b'\x1b\x4d\x01'
    data += b'\x1d\x21\x00'
    data += codigo_texto.encode()
    data += b'\n'
    salto_dots = 50  # valor calibrado — NO tocar
    data += b'\x1b\x4a' + bytes([salto_dots])

    try:
        enviar_a_impresora(data)
    except RuntimeError as e:
        return {"error": str(e)}
    except Exception as e:
        return {"error": f"No se pudo imprimir: {str(e)}"}

    return {"status": "ok"}


def _muestra_dict(m: Muestra) -> dict:
    return {
        "id":                        m.id,
        "codigo_barra":              m.codigo_barra,
        "caja_id":                   m.caja_id,
        "numero_caja":               m.numero_caja,
        "nivel":                     m.nivel,
        "codigo_utn_especie":        m.codigo_utn_especie,
        "numero_replica":            m.numero_replica,
        "numero_tubo_en_caja":       m.numero_tubo_en_caja,
        "numero_muestra_ccmbi_ogem": m.numero_muestra_ccmbi_ogem,
        "medio_cultivo":             m.medio_cultivo,
        "ubicacion_refrigerador":    m.ubicacion_refrigerador or "—",
        "especie":                   m.especie or "NO",
        "seguimiento":               m.seguimiento or "NO",
        "identificacion_taxonomica": m.identificacion_taxonomica or "—",
        "origen_muestra":            m.origen_muestra or "—",
        "codigo_para_caja":          m.codigo_para_caja or "—",
        "created_at":                m.created_at.strftime("%Y-%m-%d %H:%M"),
        "barcode_url":               f"/static/barcodes/{m.codigo_barra}.png",
    }


@app.get("/muestras/{muestra_id}/json")
def muestra_json(muestra_id: int):
    db      = SessionLocal()
    muestra = db.query(Muestra).filter(Muestra.id == muestra_id).first()
    db.close()
    if not muestra:
        raise HTTPException(status_code=404, detail="No existe")
    return _muestra_dict(muestra)


# ── SCAN / BÚSQUEDA ────────────────────────────────────────────────────────────

@app.get("/buscar/{codigo}")
def buscar_codigo(codigo: str):
    if codigo.isdigit():
        codigo = f"UTN-2026-{codigo.zfill(5)}"
    db      = SessionLocal()
    muestra = db.query(Muestra).filter(Muestra.codigo_barra == codigo).first()
    db.close()
    if not muestra:
        raise HTTPException(status_code=404, detail="Muestra no encontrada")
    return _muestra_dict(muestra)


# ── EXPORTAR (PDF + Excel) ─────────────────────────────────────────────────────

@app.get("/exportar")
def exportar(
    fmt:   str = Query(...),
    scope: str = Query("muestras"),
    ids:   str = Query(None),
):
    db = SessionLocal()
    try:
        id_list = [int(i) for i in ids.split(",") if i.strip().isdigit()] if ids else []
        if scope == "muestras":
            q = db.query(Muestra).order_by(Muestra.created_at.desc())
            if id_list:
                q = q.filter(Muestra.id.in_(id_list))
            registros = q.all()
        else:
            q = db.query(Caja).order_by(Caja.created_at.desc())
            if id_list:
                q = q.filter(Caja.id.in_(id_list))
            registros = q.all()
    finally:
        db.close()

    now_str       = datetime.now().strftime("%Y-%m-%d %H:%M")
    fecha_archivo = datetime.now().strftime("%Y%m%d_%H%M")

    # ── PDF ────────────────────────────────────────────────────────────────────
    if fmt == "pdf":
        if scope == "muestras":
            cabeceras = [
                "ID", "Código", "Especie UTN", "Caja", "Tubo", "Réplica",
                "Medio", "CCMBIOGEM", "Especie", "Seguim.",
                "Id. Tax.", "Origen", "Cód. Caja", "Ubicación", "Fecha"
            ]
            filas = [
                [f"#{m.id}", m.codigo_barra, m.codigo_utn_especie, m.numero_caja,
                 m.numero_tubo_en_caja, m.numero_replica, m.medio_cultivo,
                 m.numero_muestra_ccmbi_ogem, m.especie or "NO", m.seguimiento or "NO",
                 m.identificacion_taxonomica or "—", m.origen_muestra or "—",
                 m.codigo_para_caja or "—", m.ubicacion_refrigerador or "—",
                 m.created_at.strftime("%Y-%m-%d")]
                for m in registros
            ]
            titulo_seccion = "Registro de Muestras"
        else:
            db2 = SessionLocal()
            cabeceras = ["ID", "Nombre", "Congelador", "Piso", "Posición", "Fecha", "Muestras"]
            filas = [
                [f"#{c.id}", c.nombre,
                 "Vertical" if c.congelador == 1 else "Horizontal",
                 c.piso, c.posicion, c.fecha,
                 contar_muestras_en_caja(db2, c.id)]
                for c in registros
            ]
            db2.close()
            titulo_seccion = "Registro de Cajas"

        filas_html = "".join(
            "<tr>" + "".join(f"<td>{str(v)}</td>" for v in fila) + "</tr>"
            for fila in filas
        )
        ths         = "".join(f"<th>{h}</th>" for h in cabeceras)
        filtros_str = f"IDs seleccionados: {ids}" if ids else "Todos los registros"

        logo_b64   = get_logo_b64()
        logo_html  = (f'<img src="{logo_b64}" style="height:36px;object-fit:contain;">'
                      if logo_b64 else
                      '<span style="font-size:16px;font-weight:bold;color:#2563eb;">UTN · Laboratorio BIOGEM</span>')

        html = f"""<!doctype html><html><head><meta charset="utf-8">
        <style>
            @page {{ size: A4 landscape; margin: 15mm 12mm; }}
            body {{ font-family: Arial, sans-serif; font-size: 9px; color: #1e293b; }}
            .header {{ margin-bottom: 14px; border-bottom: 2px solid #2563eb; padding-bottom: 10px;
                       display: flex; justify-content: space-between; align-items: center; }}
            .header-left {{ display: flex; align-items: center; gap: 12px; }}
            .header-left h2 {{ font-size: 13px; font-weight: bold; color: #1e293b; margin: 0; }}
            .header-left h3 {{ font-size: 10px; font-weight: normal; color: #475569; margin: 3px 0 0; }}
            .header-right {{ text-align: right; font-size: 8px; color: #94a3b8; }}
            .meta {{ margin-bottom: 10px; font-size: 8px; color: #64748b;
                     background: #f1f5f9; padding: 5px 8px; border-radius: 4px; }}
            table {{ width: 100%; border-collapse: collapse; }}
            th {{ background: #2563eb; color: white; padding: 6px 7px; text-align: left;
                  font-size: 8px; font-weight: bold; text-transform: uppercase; }}
            td {{ padding: 5px 7px; border-bottom: 1px solid #e2e8f0; font-size: 8.5px; }}
            tr:nth-child(even) td {{ background: #f8fafc; }}
            .footer {{ margin-top: 14px; font-size: 7.5px; color: #94a3b8;
                       border-top: 1px solid #e2e8f0; padding-top: 6px;
                       display: flex; justify-content: space-between; }}
            .total {{ font-size: 9px; font-weight: bold; color: #2563eb; margin-bottom: 8px; }}
        </style></head><body>
        <div class="header">
            <div class="header-left">
                {logo_html}
                <div>
                    <h2>{titulo_seccion}</h2>
                    <h3>Sistema de Trazabilidad · BIOGEM</h3>
                </div>
            </div>
            <div class="header-right">Generado: {now_str}<br>UTN · Laboratorio BIOGEM</div>
        </div>
        <div class="meta"><b>Filtros:</b> {filtros_str}</div>
        <div class="total">Total de registros: {len(filas)}</div>
        <table><thead><tr>{ths}</tr></thead><tbody>{filas_html}</tbody></table>
        <div class="footer">
            <span>Maintronic · info@maintronic.com.ec · (593) 02 266 6256</span>
            <span>UTN · Laboratorio BIOGEM — Sistema de Trazabilidad de Muestras</span>
        </div></body></html>"""

        pdf_bytes = WeasyHTML(string=html).write_pdf()
        filename  = f"UTN_{titulo_seccion.replace(' ','_')}_{fecha_archivo}.pdf"
        return Response(
            content=pdf_bytes,
            media_type="application/pdf",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'}
        )

    # ── Excel ──────────────────────────────────────────────────────────────────
    elif fmt == "excel":
        wb = openpyxl.Workbook()
        ws = wb.active

        azul_oscuro = "1E3A8A"
        azul_medio  = "2563EB"
        azul_claro  = "DBEAFE"
        blanco      = "FFFFFF"
        gris_fila   = "F8FAFC"
        borde_color = "E2E8F0"
        thin        = Side(style="thin", color=borde_color)
        borde       = Border(left=thin, right=thin, top=thin, bottom=thin)

        if scope == "muestras":
            cabeceras = [
                "ID", "Código de barras", "Especie UTN", "Caja", "Tubo", "Réplica",
                "Medio de cultivo", "N° CCMBIOGEM", "Especie", "Seguimiento",
                "Id. Taxonómica", "Origen muestra", "Cód. para caja",
                "Ubicación", "Fecha registro"
            ]
            titulo_seccion = "Registro de Muestras"
            ws.title       = "Muestras"
        else:
            cabeceras      = ["ID", "Nombre", "Congelador", "Piso", "Posición", "Fecha", "Registrado", "Muestras"]
            titulo_seccion = "Registro de Cajas"
            ws.title       = "Cajas"

        ncols    = len(cabeceras)
        col_last = get_column_letter(ncols)

        ws.merge_cells(f"A1:{col_last}1")
        c           = ws["A1"]
        c.value     = f"UTN · Laboratorio BIOGEM — {titulo_seccion}"
        c.font      = Font(name="Arial", bold=True, size=13, color=blanco)
        c.fill      = PatternFill("solid", fgColor=azul_oscuro)
        c.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[1].height = 36

        ws.merge_cells(f"A2:{col_last}2")
        c           = ws["A2"]
        c.value     = f"Generado: {now_str}   —   Total: {len(registros)} registros"
        c.font      = Font(name="Arial", size=9, color="64748B")
        c.fill      = PatternFill("solid", fgColor="F1F5F9")
        c.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[2].height = 16
        ws.row_dimensions[3].height = 6

        for ci, h in enumerate(cabeceras, 1):
            c           = ws.cell(row=4, column=ci, value=h)
            c.font      = Font(name="Arial", bold=True, size=9, color=blanco)
            c.fill      = PatternFill("solid", fgColor=azul_medio)
            c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            c.border    = borde
        ws.row_dimensions[4].height = 22

        db2 = SessionLocal() if scope != "muestras" else None
        for ri, reg in enumerate(registros, 5):
            row_fill = PatternFill("solid", fgColor=(gris_fila if ri % 2 == 0 else blanco))
            if scope == "muestras":
                m    = reg
                vals = [
                    m.id, m.codigo_barra, m.codigo_utn_especie, m.numero_caja,
                    m.numero_tubo_en_caja, m.numero_replica, m.medio_cultivo,
                    m.numero_muestra_ccmbi_ogem, m.especie or "NO", m.seguimiento or "NO",
                    m.identificacion_taxonomica or "—", m.origen_muestra or "—",
                    m.codigo_para_caja or "—", m.ubicacion_refrigerador or "—",
                    m.created_at.strftime("%Y-%m-%d %H:%M"),
                ]
            else:
                c_obj = reg
                vals  = [
                    c_obj.id, c_obj.nombre,
                    "Vertical" if c_obj.congelador == 1 else "Horizontal",
                    c_obj.piso, c_obj.posicion, c_obj.fecha,
                    c_obj.created_at.strftime("%Y-%m-%d %H:%M"),
                    contar_muestras_en_caja(db2, c_obj.id),
                ]
            for ci, v in enumerate(vals, 1):
                cell           = ws.cell(row=ri, column=ci, value=v)
                cell.font      = Font(name="Arial", size=9)
                cell.fill      = row_fill
                cell.alignment = Alignment(vertical="center")
                cell.border    = borde
            ws.row_dimensions[ri].height = 16
        if db2:
            db2.close()

        last_row = 4 + len(registros) + 1
        ws.merge_cells(f"A{last_row}:{col_last}{last_row}")
        c           = ws.cell(row=last_row, column=1, value=f"Total: {len(registros)}")
        c.font      = Font(name="Arial", bold=True, size=9, color=azul_medio)
        c.fill      = PatternFill("solid", fgColor=azul_claro)
        c.alignment = Alignment(horizontal="right", vertical="center")
        ws.row_dimensions[last_row].height = 18

        for ci in range(1, ncols + 1):
            col_letter = get_column_letter(ci)
            max_len    = max(
                (len(str(ws.cell(row=r, column=ci).value or ""))
                 for r in range(4, last_row + 1)),
                default=8
            )
            ws.column_dimensions[col_letter].width = max(10, min(max_len + 3, 44))

        ws.freeze_panes = "A5"

        ws_info        = wb.create_sheet("Info")
        ws_info["A1"] = "Sistema de Trazabilidad UTN · Laboratorio"
        ws_info["A1"].font = Font(bold=True, size=12)
        ws_info["A3"] = "Desarrollado por:"; ws_info["B3"] = "Maintronic"
        ws_info["A4"] = "Contacto:";         ws_info["B4"] = "info@maintronic.com.ec"
        ws_info["A5"] = "Teléfono:";         ws_info["B5"] = "(593) 02 266 6256 / 09 979 6375"
        ws_info["A7"] = "Archivo generado:"; ws_info["B7"] = now_str
        ws_info["A8"] = "Sección:";          ws_info["B8"] = titulo_seccion
        ws_info["A9"] = "Total registros:";  ws_info["B9"] = len(registros)
        ws_info.column_dimensions["A"].width = 22
        ws_info.column_dimensions["B"].width = 35

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        filename = f"UTN_{titulo_seccion.replace(' ','_')}_{fecha_archivo}.xlsx"
        return Response(
            content=buffer.getvalue(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'}
        )

    raise HTTPException(status_code=400, detail="Formato no soportado. Usa fmt=pdf o fmt=excel.")


# ── PDF etiqueta térmica ───────────────────────────────────────────────────────

@app.get("/muestras/{muestra_id}/pdf")
def generar_pdf_etiqueta(muestra_id: int):
    db      = SessionLocal()
    muestra = db.query(Muestra).filter(Muestra.id == muestra_id).first()
    db.close()
    if not muestra:
        return {"error": "No existe"}

    base_path    = os.path.abspath("app/static/barcodes/")
    image_path   = f"file://{base_path}/{muestra.codigo_barra}.png"
    html_content = f"""<html><head><style>
        @page {{ size: 32mm 13mm; margin: 0; }}
        html, body {{ margin: 0; padding: 0; width: 32mm; height: 13mm; }}
        .label {{ width: 32mm; height: 13mm; display: flex; flex-direction: column;
                  justify-content: center; align-items: center; font-family: Arial; }}
        img {{ width: 30mm; }}
        .codigo {{ font-size: 8px; font-weight: bold; margin-top: 1mm; }}
    </style></head>
    <body><div class="label">
        <img src="{image_path}">
        <div class="codigo">{muestra.codigo_barra}</div>
    </div></body></html>"""
    pdf = WeasyHTML(string=html_content).write_pdf()
    return Response(content=pdf, media_type="application/pdf")


# ── CALIBRACIÓN ────────────────────────────────────────────────────────────────

@app.get("/calibrar/{salto}")
def calibrar(salto: int):
    data           = b''
    codigo_barcode = b'00001'
    codigo_texto   = b'UTN-2026-00001'
    for _ in range(5):
        data += b'\x1b\x40' + b'\x1b\x61\x02' + b'\x1d\x68\x28' + b'\x1d\x77\x02' + b'\x1d\x48\x00'
        data += b'\x1d\x6b\x49' + bytes([len(codigo_barcode)]) + codigo_barcode
        data += b'\x1b\x4a\x04' + b'\x1b\x4d\x01' + b'\x1d\x21\x00' + codigo_texto + b'\n'
        data += b'\x1b\x4a' + bytes([salto])
    try:
        enviar_a_impresora(data)
    except Exception as e:
        return {"error": str(e)}
    return {"status": f"Impreso con salto={salto}"}